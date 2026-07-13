"""
build.py — Script build tự động cho Dashboard Theo dõi mốc nhập nhận xét (Scots English)

Cách dùng:
    python3 build.py

Input (đặt trong thư mục data/, ĐÚNG TÊN FILE như bên dưới — ghi đè file cũ mỗi lần cập nhật):
    data/class_info.xlsx      <- file "All_Branch_Class_Info" export từ LMS (sheet "Class Info")
    data/student_issue.xlsx   <- file "LMS_Student_Issue" export từ LMS (sheet "Student Issue")
    data/grade_report.xlsx    <- file "BC điểm CK" export từ LMS (sheet "Export")
    data/holidays.xlsx        <- file "Ngày nghỉ lễ" của công ty (cột A = ngày nghỉ, sheet "Export")

    Cả 4 file đều KHÔNG bắt buộc phải có mặt — nếu thiếu file nào, script vẫn chạy,
    chỉ là phần dữ liệu tương ứng sẽ không có (riêng thiếu holidays.xlsx sẽ dùng danh sách
    dự phòng có sẵn trong code, xem HOLIDAYS_FALLBACK bên dưới).

Output:
    index.html   <- file dashboard hoàn chỉnh, deploy qua GitHub Pages

Muốn cập nhật lịch nghỉ lễ năm mới: CHỈ CẦN thay file data/holidays.xlsx bằng file mới của công ty,
không cần sửa code (script tự đọc lại từ file mỗi lần build).
Muốn thêm/sửa mapping Vùng - Chi nhánh - AS: sửa biến REGION_MAPPING_RAW bên dưới.
"""

import json
import re
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime, timedelta

import openpyxl

# =====================================================================================
# CẤU HÌNH CHUNG
# =====================================================================================

MILESTONES = [6, 12, 25, 36, 50]
MILESTONE_LABELS = {6: "Buổi 6", 12: "Buổi 12", 25: "Giữa kỳ (25)", 36: "Buổi 36", 50: "Cuối kỳ (50)"}
LABEL_TO_NUM = {v: k for k, v in MILESTONE_LABELS.items()}

DATA_DIR = "data"
CLASS_INFO_PATH = f"{DATA_DIR}/class_info.xlsx"
STUDENT_ISSUE_PATH = f"{DATA_DIR}/student_issue.xlsx"
GRADE_REPORT_PATH = f"{DATA_DIR}/grade_report.xlsx"
HOLIDAYS_PATH = f"{DATA_DIR}/holidays.xlsx"
TEMPLATE_PATH = "template.html"
OUTPUT_PATH = "index.html"


# =====================================================================================
# 1. MAPPING VÙNG - CHI NHÁNH - AS
#    Cập nhật ở đây khi có chi nhánh mới hoặc đổi AS phụ trách.
#    Cột 3 (AS) không dùng để tính toán gì, chỉ hiển thị & lọc.
# =====================================================================================

REGION_MAPPING_RAW = """Vùng 1\tKim Giang\tNguyễn Thị Thủy
Vùng 1\tThanh Hóa\tVũ Thị Nga
Vùng 1\tLam Sơn\tVũ Thị Nga
Vùng 1\tNguyễn Xiển\tNguyễn Thị Hường
Vùng 1\tLinh Đàm\tVũ Như Bình
Vùng 1\tTây Hồ\tBùi Huyền Diễm
Vùng 1\tNguyễn Tuân\tHồ Thị Hương Giang
Vùng 1\tHoàng Đạo Thúy\tLê Thị Thanh Tâm
Vùng 1\tHoàng Quốc Việt\tBùi Huyền Diễm
Vùng 1\tTrung Văn\tLê Thị Thanh Tâm
Vùng của Liên\tHải Dương\tMa Công Quỳnh
Vùng của Liên\tVĩnh Phúc\tVũ Ngọc Ý
Vùng của Liên\tLong Biên\tLê Thị Hà
Vùng của Liên\tVĩnh Phúc 3\tBùi Văn Linh
Vùng của Liên\tPhúc Yên\tBùi Thị Thanh Lãng
Vùng của Liên\tViệt Trì\tĐỗ Thảo Vân
Vùng của Liên\tMỹ Đình\tNguyễn Thị Mai Anh 1
Vùng của Liên\tVinhomes Gardenia\tNguyễn Thị Thương
Vùng 3\tTimes City\tTriệu Thị Bảo Ngân
Vùng 3\tVăn Khê\tNguyễn Thị Mỹ Linh
Vùng 3\tAn Khánh\tLê Thị Thu Trang
Vùng 3\tVinhomes Smart City\tLưu Thu Phương
Vùng 3\tVinhomes Smart City 2\tNguyễn Thị Ngọc
Vùng 3\tDương Nội\tMai Thị Tùng Anh / Vũ Thị Thanh / Lê Thị Thanh
Vùng 3\tPhạm Văn Đồng\tNguyễn Công Thùy Linh
Vùng 3\tThái Bình\tĐào Thị Huệ
Vùng 5\tTừ Sơn\tNguyễn Thị Mai Anh
Vùng 5\tHải Phòng 2\tPhạm Thị Thuận
Vùng 5\tHải Phòng\tBùi Thị Bích Phương
Vùng 5\tBắc Ninh\tĐặng Thị Thúy
Vùng 5\tBắc Ninh 2\tLý Thị Lan
Vùng 5\tBắc Giang\tNguyễn Thị Thùy Trang
Vùng 6\tSài Đồng\tLê Lâm Hải
Vùng 6\tVinh\tNguyễn Thị Ngọc
Vùng 6\tOcean Park\tPhạm Thị Dinh
Vùng 6\tTrường Chinh\tHoàng Liên Nhi
Vùng 6\tĐịnh Công\tVũ Hồng Nhung
Vùng 7\tĐà Nẵng\tNguyễn Thị Trang
Vùng 7\tĐà Nẵng 2\tNguyễn Thị Trang
Vùng 7\tPhan Văn Trị\tNguyễn Ngọc Minh Thư
Vùng 8\tCeladon - Tân Phú\tPhùng Hà Kiều Anh
Vùng 8\tPhạm Văn Chiêu\tPhùng Hà Kiều Anh
Vùng 8\tGrand Park\tPhùng Hà Kiều Anh"""


def build_region_map():
    region_map = {}
    for line in REGION_MAPPING_RAW.strip().split("\n"):
        region, campus, asn = line.split("\t")
        branch_full = "Scots English " + campus.strip()
        region_map[branch_full] = {"region": region.strip(), "as": asn.strip()}
    return region_map


# =====================================================================================
# 2. LỊCH NGHỈ LỄ CHÍNH THỨC VIỆT NAM
#    Thêm năm mới vào đây khi công ty công bố lịch nghỉ chính thức của năm đó
#    (nguồn: file "Ngày nghỉ lễ.xlsx" do công ty cung cấp — không phải lịch nghỉ chung
#    của nhà nước, vì công ty có thể nghỉ ít/nhiều hơn tuỳ năm).
# =====================================================================================

# Danh sách dự phòng (dùng khi KHÔNG có file data/holidays.xlsx) — chính xác đến ngày build.py này
# được viết. Khuyến khích luôn cập nhật qua file data/holidays.xlsx thay vì sửa danh sách này.
HOLIDAYS_FALLBACK = [
    "2023-01-01", "2023-04-29", "2023-04-30", "2023-05-01", "2023-05-02", "2023-05-03",
    "2023-09-01", "2023-09-02", "2023-09-03", "2023-09-04",
    "2024-01-01", "2024-02-08", "2024-02-09", "2024-02-10", "2024-02-11", "2024-02-12",
    "2024-02-13", "2024-02-14", "2024-04-30", "2024-05-01", "2024-05-02", "2024-05-03", "2024-09-01",
    "2025-01-01", "2025-01-25", "2025-01-26", "2025-01-27", "2025-01-28", "2025-01-29",
    "2025-01-30", "2025-01-31", "2025-02-01", "2025-02-02", "2025-04-07", "2025-04-30",
    "2025-05-01", "2025-08-30", "2025-08-31", "2025-09-01", "2025-09-02",
    "2026-01-01", "2026-02-14", "2026-02-15", "2026-02-16", "2026-02-17", "2026-02-18",
    "2026-02-19", "2026-02-20", "2026-02-21", "2026-02-22", "2026-02-23", "2026-04-26",
    "2026-04-27", "2026-04-30", "2026-05-01", "2026-05-02", "2026-05-03", "2026-09-01", "2026-09-02",
]


def build_vn_holidays():
    """Đọc ngày nghỉ lễ từ data/holidays.xlsx (cột A, sheet 'Export' hoặc sheet đầu tiên).
    Nếu không tìm thấy file, dùng HOLIDAYS_FALLBACK để dashboard vẫn build được."""
    try:
        wb = openpyxl.load_workbook(HOLIDAYS_PATH, data_only=True)
    except FileNotFoundError:
        print(f"[CẢNH BÁO] Không tìm thấy {HOLIDAYS_PATH} — dùng danh sách ngày nghỉ dự phòng có sẵn trong code.")
        return set(datetime.strptime(d, "%Y-%m-%d").date() for d in HOLIDAYS_FALLBACK)

    ws = wb["Export"] if "Export" in wb.sheetnames else wb[wb.sheetnames[0]]
    holidays = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0] if row else None
        if val is None:
            continue
        if hasattr(val, "date"):
            holidays.add(val.date())
        elif isinstance(val, str):
            try:
                holidays.add(datetime.strptime(val.strip(), "%Y-%m-%d").date())
            except ValueError:
                continue  # bỏ qua dòng không parse được (VD dòng "No filters applied" ở cuối file)
    print(f"[holidays] Đọc được {len(holidays)} ngày nghỉ lễ từ {HOLIDAYS_PATH}.")
    return holidays


VN_HOLIDAYS = build_vn_holidays()
WEEKDAY_MAP = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6}


def nearest_milestone(n):
    return min(MILESTONES, key=lambda m: abs(m - n))


# =====================================================================================
# 3. ĐỌC FILE CLASS INFO -> DANH SÁCH LỚP (classes_v2)
# =====================================================================================

def load_class_info(path, region_map):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"[CẢNH BÁO] Không tìm thấy {path} — bỏ qua, dashboard sẽ không có dữ liệu lớp.")
        return [], {}
    ws = wb["Class Info"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    data = []
    class_schedule = {}
    skipped = 0
    for r in rows:
        branch, program, class_name, start_date, end_date, capacity, teacher_id, teacher_name, days, completed = r
        if teacher_name is None or completed is None:
            skipped += 1
            continue
        cur, total = capacity.split("/") if capacity else ("0", "0")
        info = region_map.get(branch)
        row = {
            "branch": branch,
            "program": program,
            "class_name": class_name,
            "start_date": start_date if isinstance(start_date, str) else str(start_date),
            "end_date": end_date if isinstance(end_date, str) else str(end_date),
            "cap_cur": int(cur),
            "cap_total": int(total),
            "teacher": teacher_name,
            "days": days,
            "completed": int(completed),  # giữ lại để đối chiếu (tooltip); tiến độ hiển thị tự tính lại trong JS
            "region": info["region"] if info else "Chưa xác định",
            "as": info["as"] if info else "",
        }
        data.append(row)
        if class_name and row["start_date"] and days:
            class_schedule[class_name] = {
                "start_date": row["start_date"],
                "days": [d.strip() for d in days.split(",")],
            }

    print(f"[class_info] {len(data)} lớp hợp lệ, bỏ qua {skipped} lớp rỗng (chưa có GV/học viên).")
    unmapped_branches = sorted(set(d["branch"] for d in data if d["region"] == "Chưa xác định"))
    if unmapped_branches:
        print(f"[class_info] CẢNH BÁO: {len(unmapped_branches)} chi nhánh chưa có trong mapping Vùng: {unmapped_branches}")
    return data, class_schedule


def estimate_lesson_number(class_name, comment_date, class_schedule):
    info = class_schedule.get(class_name)
    if not info or not comment_date:
        return None
    try:
        start = datetime.strptime(info["start_date"], "%Y-%m-%d").date()
    except Exception:
        return None
    if isinstance(comment_date, datetime):
        target = comment_date.date()
    elif isinstance(comment_date, str):
        try:
            target = datetime.strptime(comment_date, "%Y-%m-%d").date()
        except Exception:
            return None
    else:
        target = comment_date
    if target < start:
        return None
    target_wds = set(WEEKDAY_MAP[d] for d in info["days"] if d in WEEKDAY_MAP)
    if not target_wds:
        return None
    count = 0
    d = start
    while d <= target:
        if d.weekday() in target_wds and d not in VN_HOLIDAYS:
            count += 1
        d += timedelta(days=1)
    return count


# =====================================================================================
# 4. PHÂN LOẠI NHẬN XÉT TỪ FILE STUDENT ISSUE THEO MỐC (6/12/25/36/50)
# =====================================================================================

RELEVANT_TYPES = {
    "Instructor Comment for Consultation",
    "Learning Portal Comment",
    "Grades",
    "Grade/Class/Level",
    "Others",
}

LESSON_KW = r"(?:lesson|session|bu[oổ]i|buoi)"
NUM_NEAR_KW_PATTERNS = [
    re.compile(rf"(\d{{1,2}})\s*(?:st|nd|rd|th)?\s*{LESSON_KW}", re.IGNORECASE),
    re.compile(rf"{LESSON_KW}\s*[:#\-]?\s*(\d{{1,2}})", re.IGNORECASE),
    re.compile(r"\bl\s*(\d{1,2})\b", re.IGNORECASE),
]
MIDTERM_PAT = re.compile(r"\bmid[\s\-]?term\b|\bmed[\s\-]?term\b|\bm\s*t\b", re.IGNORECASE)
FINAL_PAT = re.compile(r"\bfinal\b", re.IGNORECASE)
PROGRESS_IDX_PAT = re.compile(r"progress\D{0,10}?(\d)\b", re.IGNORECASE)
# Đã hiệu chỉnh theo dữ liệu thực tế (đối chiếu với ngày + lịch học): PR1≈buổi12, PR2≈Giữa kỳ, PR3≈buổi36
PROGRESS_TO_MILESTONE = {"1": 12, "2": 25, "3": 36}


def classify_subject(subject, class_code):
    if not subject:
        return None, None
    text = subject
    if class_code and class_code in text:
        text = text.replace(class_code, " ")
    for pat in NUM_NEAR_KW_PATTERNS:
        m = pat.search(text)
        if m:
            n = int(m.group(1))
            if 1 <= n <= 60:
                milestone = nearest_milestone(n)
                conf = "chắc chắn" if abs(milestone - n) <= 2 else "gần đúng"
                return milestone, f"explicit_number({n})_{conf}"
    if MIDTERM_PAT.search(text):
        return 25, "keyword_midterm"
    if FINAL_PAT.search(text):
        return 50, "keyword_final"
    m = PROGRESS_IDX_PAT.search(text)
    if m and m.group(1) in PROGRESS_TO_MILESTONE:
        return PROGRESS_TO_MILESTONE[m.group(1)], f"progress_index_calibrated({m.group(1)})"
    return None, None


def classify_row(row, class_schedule):
    milestone, method = classify_subject(row.get("Subject"), row.get("Class"))
    if milestone:
        return milestone, method
    est = estimate_lesson_number(row.get("Class"), row.get("Date"), class_schedule)
    if est is not None:
        milestone = nearest_milestone(est)
        diff = abs(milestone - est)
        if diff <= 3:
            return milestone, f"date_estimate({est})"
        return None, f"date_estimate_out_of_range({est})"
    return None, "unclassified"


def confidence_of(method):
    if method is None:
        return "unknown"
    if method.startswith("explicit_number") and "chắc chắn" in method:
        return "cao"
    if method.startswith("explicit_number"):
        return "trung_binh"
    if method.startswith("keyword_midterm") or method.startswith("keyword_final"):
        return "cao"
    if method.startswith("progress_index_calibrated"):
        return "trung_binh"
    if method.startswith("date_estimate") and "out_of_range" not in method:
        return "trung_binh"
    return "unknown"


def load_student_issue(path, class_schedule):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"[CẢNH BÁO] Không tìm thấy {path} — bỏ qua, sẽ không có nhận xét tự động từ Student Issue.")
        return {}
    ws = wb["Student Issue"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = ["No", "Branch", "Date", "Std.Name", "Nick Name", "ID", "Grade", "Class",
              "A.Staff", "Type", "Status", "Creator", "Subject", "Link", "Detail"]

    relevant = []
    for r in rows:
        d = dict(zip(header, r))
        if d["Type"] in RELEVANT_TYPES:
            relevant.append(d)

    TRUNCATE = 260
    index = defaultdict(lambda: defaultdict(list))
    method_counter = Counter()
    unresolved = 0
    for r in relevant:
        milestone, method = classify_row(r, class_schedule)
        method_counter[(method or "unclassified").split("(")[0]] += 1
        if milestone is None:
            unresolved += 1
            continue
        cls = r.get("Class")
        if not cls:
            continue
        detail = (r.get("Detail") or "").strip()
        truncated = len(detail) > TRUNCATE
        entry = {
            "student": r.get("Std.Name") or "",
            "date": (r.get("Date").strftime("%Y-%m-%d") if hasattr(r.get("Date"), "strftime") else str(r.get("Date") or ""))[:10],
            "detail": detail[:TRUNCATE],
            "truncated": truncated,
            "conf": confidence_of(method),
            "source": "student_issue",
            "type": r.get("Type") or "",
        }
        index[cls][str(milestone)].append(entry)

    print(f"[student_issue] {len(relevant)} dòng liên quan, {unresolved} dòng không xác định được mốc ({unresolved/len(relevant)*100:.1f}%).")
    print(f"[student_issue] Phương pháp phân loại: {dict(method_counter)}")

    final_index = {}
    for cls, milestones in index.items():
        final_index[cls] = {}
        for m, entries in milestones.items():
            has_af = any(e.get("type") == "Learning Portal Comment" for e in entries)
            has_teacher = any(e.get("type") != "Learning Portal Comment" for e in entries)
            final_index[cls][m] = {"count": len(entries), "items": entries, "has_af": has_af, "has_teacher": has_teacher}
    return final_index


# =====================================================================================
# 5. BẢNG ĐIỂM CK (MIDTERM/FINAL) — TEACHER COMMENT + AF_COMMENT
# =====================================================================================

def detect_grade_milestone(report_name):
    if not report_name:
        return None
    name = report_name.upper()
    if "MIDTERM" in name:
        return 25
    if "FINAL" in name:
        return 50
    return None


def load_grade_report(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"[CẢNH BÁO] Không tìm thấy {path} — bỏ qua, sẽ không có nhận xét từ bảng điểm CK.")
        return {}
    ws = wb["Export"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    TRUNCATE = 260
    new_index = defaultdict(lambda: defaultdict(list))
    skipped_no_comment = 0
    for r in rows:
        (mien, vung, chi_nhanh, ma_lop, ten_lop, ma_hv, ten_hv, ten_bang_diem, ngay_test,
         listening, reading, speaking, writing, overall, hw, interaction, academic, total,
         teacher_comment, af_comment) = r

        if not ma_lop:
            continue
        milestone = detect_grade_milestone(ten_bang_diem)
        if milestone is None:
            continue
        comment = (teacher_comment or "").strip()
        if not comment:
            skipped_no_comment += 1
            continue

        truncated = len(comment) > TRUNCATE
        entry = {
            "student": ten_hv or "",
            "date": ngay_test.strftime("%Y-%m-%d") if hasattr(ngay_test, "strftime") else str(ngay_test or ""),
            "detail": comment[:TRUNCATE],
            "truncated": truncated,
            "conf": "cao",
            "source": "grade_report",
            "report_name": ten_bang_diem,
        }
        if af_comment and str(af_comment).strip():
            af = str(af_comment).strip()
            entry["af_comment"] = af[:TRUNCATE]
            entry["af_truncated"] = len(af) > TRUNCATE

        new_index[ma_lop][str(milestone)].append(entry)

    new_count = sum(len(v) for cls in new_index.values() for v in cls.values())
    print(f"[grade_report] {new_count} nhận xét (Teacher comment) trên {len(new_index)} lớp, "
          f"bỏ qua {skipped_no_comment} dòng không có Teacher comment.")
    return new_index


def merge_grade_into_index(comments_index, grade_index):
    for cls, milestones in grade_index.items():
        if cls not in comments_index:
            comments_index[cls] = {}
        for m, items in milestones.items():
            if m not in comments_index[cls]:
                comments_index[cls][m] = {"count": 0, "items": [], "has_af": False, "has_teacher": False}
            comments_index[cls][m]["items"].extend(items)
            comments_index[cls][m]["count"] = len(comments_index[cls][m]["items"])

    # tính lại has_af / has_teacher cho MỌI (lớp, mốc) sau khi đã gộp đầy đủ 2 nguồn
    for cls, milestones in comments_index.items():
        for m, obj in milestones.items():
            items = obj["items"]
            obj["has_af"] = any(
                it.get("type") == "Learning Portal Comment" or (it.get("source") == "grade_report" and it.get("af_comment"))
                for it in items
            )
            obj["has_teacher"] = any(it.get("type") != "Learning Portal Comment" for it in items)
    return comments_index


# =====================================================================================
# 6. MAIN — CHẠY TOÀN BỘ PIPELINE VÀ XUẤT index.html
# =====================================================================================

def main():
    region_map = build_region_map()

    classes, class_schedule = load_class_info(CLASS_INFO_PATH, region_map)
    comments_index = load_student_issue(STUDENT_ISSUE_PATH, class_schedule)
    grade_index = load_grade_report(GRADE_REPORT_PATH)
    comments_index = merge_grade_into_index(comments_index, grade_index)

    print(f"[TỔNG] {len(classes)} lớp, {len(comments_index)} lớp có ít nhất 1 nhận xét được xác nhận.")

    with open(TEMPLATE_PATH, encoding="utf-8") as f:
        template = f.read()

    classes_json = json.dumps(classes, ensure_ascii=False)
    comments_json = json.dumps(comments_index, ensure_ascii=False, separators=(",", ":"))
    holidays_json = json.dumps(sorted(d.isoformat() for d in VN_HOLIDAYS), ensure_ascii=False)

    output = (template
              .replace("__DATA_JSON__", classes_json)
              .replace("__COMMENTS_JSON__", comments_json)
              .replace("__HOLIDAYS_JSON__", holidays_json))

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"[XONG] Đã ghi {OUTPUT_PATH} ({len(output):,} ký tự).")


if __name__ == "__main__":
    main()
